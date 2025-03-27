from openpyxl import load_workbook
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule, Rule

wb = load_workbook(filename='E:\ETL\Siddhesh\Finance_Recon.xlsx')
sheet = wb.active

redFill = PatternFill(start_color='EE1111',
                      end_color='EE1111',
                      fill_type='solid')
greenFill = PatternFill(start_color='00339966',
                      end_color='00339966',
                      fill_type='solid')
yellowFill = PatternFill(start_color='00FF9900',
                      end_color='00FF9900',
                      fill_type='solid')
NoFill = PatternFill(start_color='00FFFFFF',
                      end_color='00FFFFFF',
                      fill_type='solid')

column_index = 7

for row in sheet.iter_rows(min_col=column_index, max_col=column_index):
    for cell in row:
        if cell.value is None:
            cell.fill = NoFill
        else:
            try:
                numeric_value = float(cell.value)
                if numeric_value > 0:
                    cell.fill = yellowFill
                elif numeric_value < 0:
                    cell.fill = redFill
                else:
                    cell.fill = greenFill
            except ValueError:
                cell.fill = NoFill
wb.save("E:\ETL\Siddhesh\Finance_Recon.xlsx")